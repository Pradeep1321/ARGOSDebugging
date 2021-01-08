from P4 import P4,P4Exception    # Import the module
import pprint
import csv
import openpyxl
import sys
import xml.sax
import xmltodict
import collections
import xlwings
import gc
import pandas

p4Params = {'Port': "lv-perforce03.ad.agi:1668", \
            'Pass': "Jatin#$2", \
            'User': "pnimmanapalli", \
            'Client': "myclient"}

ExcelSheet = "C:\Share\Games.xlsx"
tempFile = "C:\Share\Temp.xml"


class GroupHandler(xml.sax.ContentHandler):
    def __init__(self):
        self.data = ''
    def startElement(self,name,attrs):
        self.current = name
        if attrs._attrs != {}:
            self.data= self.data+ "\n" + name + " = " + str(attrs._attrs)
    def characters(self, content):
        self.text = content

    def endElement(self, name):
        if self.current == name and self.text.strip() != '':
            self.data = self.data + "\n"+ name + " : " + self.text
    def datatobeReturned(self):
        return self.data

'''
def ExcelUpdate(r,c,values):
    wb = xlwings.Book(ExcelSheet)
    app = xlwings.apps.active
    ws = wb.sheets[0]
    oldval = ws.cells(r,c).value
    if oldval != None:
        ws.cells(r, c).value = oldval + "\n" + values
    else:
        ws.cells(r, c).value = values
    wb.save(ExcelSheet)
    app.quit()
'''
def ExcelUpdate(r,c,values):
    book = openpyxl.load_workbook(ExcelSheet)
    ws = book.worksheets[0]
    oldval = ws.cell(r,c).value
    if oldval != None:
        ws.cell(r, c).value = oldval + "\n" + values
    else:
        ws.cell(r, c).value = values
    book.save(ExcelSheet)





'''
def printDict(d,row,col):
    for k, v in d.items():
        print(k)
        if type(v) is collections.OrderedDict:
            printDict(v)
        elif type(v) is list:
            for j in v:
                if type(j) is collections.OrderedDict:
                    printDict(j)
        else:
            print("{0} : {1}".format(k, v))
            value = value + "\n" + k + " : " + v
    ExcelUpdate(row, col, value)
'''



try:                             # Catch exceptions with try/except
  p4 = P4(client=p4Params['Client'], port=p4Params['Port'], password=p4Params['Pass'],user = p4Params['User'])
  p4.connect()                   # Connect to the Perforce server
  p4.run_login()   # Login with the login details
  GamesDir = (p4.run("dirs", "//games/" + '*'))
  Files = ['build*.json','PresentationConfig*.xml','image_param*.txt','StatsConfig*.xml','themeConfig*.cfg','paytableconfig*.xml','progConfig*.xml','betUnitConfig*.xml','SupportedFeatures.cfg']
  GameName = []
  gamestats = []
  progConfig = []
  themeConfig = []
  paytableConfig = []
  PresentationConfig = []
  build = []
  image_param = []
  betUnitConfig = []
  DepotFiles = []
  #print(p4.run("files", "//games/007_CasioRoyale/....xml"))
  #print(p4.run("files", "//games/007_CasioRoyale/....cfg"))
  rw = 1
  for games in GamesDir:
    try:
        GameNames  = (games['dir'].split("//games/")[1])
        rw += 1
        col = 1
        ExcelUpdate(rw,col,GameNames)
        for File in Files:
            col+=1
            try:
                Depot = p4.run("files", games['dir'] + "/..."+ File)
            except:
                print("Unexpected error:", sys.exc_info()[0])
            print(Depot)
            if Depot != None:
                for i in Depot:
                    p4.run_print('-q', '-o',tempFile,i['depotFile'])
                    data = ''
                    if File.find(".xml") != -1 or File.find(".cfg") != -1:
                        handler = GroupHandler()
                        parser = xml.sax.make_parser()
                        parser.setContentHandler(handler)
                        parser.parse(tempFile)
                        data = handler.datatobeReturned()
                        del handler
                    else:
                        tf = open(tempFile,"r")
                        dataList = tf.readlines()
                        if File == "build*.json":
                            data = "".join([str(elem) for elem in dataList if elem.find("ModulePartNumber") != -1])
                        else:
                            data = "".join([str(elem) for elem in dataList])
                        tf.close()
                if data != '':
                    ExcelUpdate(rw, col, data)
                del data
                gc.collect()

    except :
        print("Unexpected error:", sys.exc_info()[0])
    '''    try:
      gamestats.append(p4.run("files", games['dir'] + "/...StatsConfig*.xml"))
    except:
        print("Unexpected error:", sys.exc_info()[0])
    try:
        progConfig.append(p4.run("files", games['dir'] + "/...progConfig*.xml")) #multiple values -
    except:
        print("Unexpected error:", sys.exc_info()[0])
    try:
        themeConfig.append(p4.run("files", games['dir'] + "/...themeConfig*.cfg"))
    except:
        print("Unexpected error:", sys.exc_info()[0])
    try:
        paytableConfig.append(p4.run("files", games['dir'] + "/...paytableconfig*.xml"))
    except:
        print("Unexpected error:", sys.exc_info()[0])
    try:
        PresentationConfig.append(p4.run("files", games['dir'] + "/...PresentationConfig*.xml"))
    except:
        print("Unexpected error:", sys.exc_info()[0])
    try:
        build.append(p4.run("files", games['dir'] + "/...build*.json"))
    except:
        print("Unexpected error:", sys.exc_info()[0])
    try:
        image_param.append(p4.run("files", games['dir'] + "/...image_param*.txt"))
    except:
        print("Unexpected error:", sys.exc_info()[0])
    try:
        betUnitConfig.append(p4.run("files", games['dir'] + "/...betUnitConfig*.txt"))
    except:
        print("Unexpected error:", sys.exc_info()[0])'''

  p4.disconnect()                # Disconnect from the server
except P4Exception:
  for e in p4.errors:            # Display errors
      p4.disconnect()  # Disconnect from the server
      print(e)